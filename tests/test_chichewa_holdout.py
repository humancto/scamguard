from pathlib import Path

from openpyxl import Workbook

from scripts.build_chichewa_holdout import read_chichewa


def write_fixture(path: Path) -> None:
    workbook = Workbook()
    original = workbook.active
    original.title = "D_CHI"
    original.append(["ID", "Text", "Label"])
    original.append(
        ["MSG1", "Tumizani ndalama ku 0999994330 pa https://pay.example/now", "fraud"]
    )
    original.append(["MSG2", "Balansi yanu ndi MK500.", "normal"])
    original.append([None, None, None])

    telco = workbook.create_sheet("telcoSMS_CHI")
    telco.append(["ID", "Text", "Label"])
    telco.append(["TELCO1", "Phukusi lanu la data latha.", "normal"])

    translation = workbook.create_sheet("D_HT")
    translation.append(["ID", "Text", "Label"])
    translation.append(["MSG1", "Send money to this number", "fraud"])
    workbook.save(path)


def test_chichewa_reader_excludes_translations_and_normalizes_sensitive_values(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    write_fixture(source)

    rows = list(read_chichewa(source))

    assert len(rows) == 3
    assert [row["label"] for row in rows] == ["SCAM", "SAFE", "SAFE"]
    assert all(row["source_language"] == "Chichewa" for row in rows)
    assert all(row["split"] == "ood" for row in rows)
    assert "<PHONE_NUMBER>" in str(rows[0]["text"])
    assert "<URL>" in str(rows[0]["text"])
    assert "0999994330" not in str(rows[0]["text"])
    assert "pay.example" not in str(rows[0]["text"])
    assert all("Send money to this number" not in str(row["text"]) for row in rows)
