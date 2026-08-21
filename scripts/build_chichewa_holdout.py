#!/usr/bin/env python3
"""Build a privacy-normalized, family-collapsed Chichewa external diagnostic."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

try:
    from scripts.build_dataset import (
        URL_RE,
        cluster_near_duplicates,
        deduplicate,
        make_row,
        normalized,
        read_jsonl,
        remove_near_overlaps,
        write_jsonl,
    )
except ModuleNotFoundError:  # Direct execution places scripts/ rather than the repo on sys.path.
    from build_dataset import (  # type: ignore[no-redef]
        URL_RE,
        cluster_near_duplicates,
        deduplicate,
        make_row,
        normalized,
        read_jsonl,
        remove_near_overlaps,
        write_jsonl,
    )

SOURCE_SHA256 = "4f83cfaab196f8fab3bdbf9c89e15313ddaa889da066335fcc2f35cc6b3f487a"
PHONE_OR_LONG_DIGITS = re.compile(
    r"(?<![A-Za-z0-9])\+?\d(?:[\d ()-]{5,}\d)(?![A-Za-z0-9])|\d{10,}"
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_chichewa(path: Path) -> Iterable[dict[str, object]]:
    """Read original-language sheets only; translations would triple-count each message."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = (
        ("D_CHI", "mixed_crowdsourced_telco_augmented_unmarked"),
        ("telcoSMS_CHI", "publisher_legitimate_telco_sheet"),
    )
    try:
        for sheet_name, provenance in sheets:
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows)
            if tuple(header[:3]) != ("ID", "Text", "Label"):
                raise ValueError(f"unexpected {sheet_name} header: {header!r}")
            for values in rows:
                source_id, raw_text, raw_label = (*values, None, None, None)[:3]
                if not source_id or not raw_text or not raw_label:
                    continue
                source_label = str(raw_label).strip().casefold()
                if source_label not in {"fraud", "normal"}:
                    raise ValueError(f"unexpected {sheet_name} label: {source_label!r}")
                label = "SCAM" if source_label == "fraud" else "SAFE"
                text = URL_RE.sub("<URL>", str(raw_text))
                row = make_row(
                    text=text,
                    label=label,
                    source="chichewa_sms_fraud",
                    source_label=source_label,
                    license_name="CC-BY-4.0",
                )
                if row is None:
                    continue
                row.update(
                    {
                        "category": "UNKNOWN" if label == "SCAM" else "NONE",
                        "split": "ood",
                        "source_language": "Chichewa",
                        "source_record_id": str(source_id),
                        "source_provenance": provenance,
                        "label_policy": (
                            "publisher_curated_fraud"
                            if label == "SCAM"
                            else "publisher_legitimate_message"
                        ),
                    }
                )
                yield row
    finally:
        workbook.close()


def reference_rows(data: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for split in (
        "train",
        "dev",
        "test",
        "ood_financial",
        "ood_wspr",
        "forum_validation",
        "ood_forum",
        "ood_azsc",
    ):
        rows.extend(read_jsonl(data / f"{split}.jsonl"))
    return rows


def build(source: Path, data: Path, output: Path) -> dict[str, object]:
    if sha256(source) != SOURCE_SHA256:
        raise ValueError("Chichewa source hash differs from the pinned publisher artifact")

    source_rows = list(read_chichewa(source))
    exact_rows, exact_dropped, exact_conflicts = deduplicate(source_rows)
    clustered, near_conflicts, near_stats = cluster_near_duplicates(exact_rows)
    representatives: dict[str, dict[str, object]] = {}
    for row in clustered:
        family_id = str(row["family_id"])
        candidate = row | {"split": "ood"}
        current = representatives.get(family_id)
        if current is None or str(candidate["id"]) < str(current["id"]):
            representatives[family_id] = candidate

    candidates = list(representatives.values())
    references = reference_rows(data)
    reference_keys = {normalized(str(row["text"])) for row in references}
    candidates = [
        row for row in candidates if normalized(str(row["text"])) not in reference_keys
    ]
    exact_overlaps_removed = len(representatives) - len(candidates)
    candidates, near_overlaps_removed = remove_near_overlaps(candidates, references)
    candidates.sort(key=lambda row: str(row["id"]))

    ids = [str(row["id"]) for row in candidates]
    families = [str(row["family_id"]) for row in candidates]
    if len(ids) != len(set(ids)) or len(families) != len(set(families)):
        raise ValueError("Chichewa diagnostic is not one-row-per-id-and-family")
    if {str(row["label"]) for row in candidates} != {"SAFE", "SCAM"}:
        raise ValueError("Chichewa diagnostic must retain both SAFE and SCAM rows")
    if any(PHONE_OR_LONG_DIGITS.search(str(row["text"])) for row in candidates):
        raise ValueError("Chichewa diagnostic retains a phone/account-like value")
    if any(URL_RE.search(str(row["text"])) for row in candidates):
        raise ValueError("Chichewa diagnostic retains a live-looking URL")

    output.mkdir(parents=True, exist_ok=True)
    artifact = output / "ood_chichewa.jsonl"
    write_jsonl(artifact, candidates)
    write_jsonl(output / "quarantine_label_conflicts.jsonl", exact_conflicts + near_conflicts)
    manifest: dict[str, object] = {
        "diagnostic_schema_version": 1,
        "purpose": (
            "post-schema-v9 multilingual diagnostic; excluded from fitting and thresholding; "
            "may inform candidate selection"
        ),
        "source": {
            "doi": "https://doi.org/10.5281/zenodo.14607454",
            "zenodo_record": 14607454,
            "metadata_revision": 12,
            "license": "CC-BY-4.0",
            "raw_sha256": SOURCE_SHA256,
            "sheets_used": ["D_CHI", "telcoSMS_CHI"],
            "translations_excluded": ["D_HT", "D_MT", "telcoSMS_HT", "telcoSMS_MT"],
            "provenance_limit": (
                "D_CHI mixes collected and augmented messages without row-level provenance"
            ),
        },
        "policy": {
            "used_for_fitting": False,
            "used_for_threshold": False,
            "privacy_normalization": "URLs and phone/account-like values replaced",
            "near_template_hamming_max": 6,
            "one_representative_per_family": True,
            "native_language_review_complete": False,
            "real_row_count_claim_allowed": False,
        },
        "counts": {
            "source_rows": len(source_rows),
            "source_labels": dict(Counter(str(row["label"]) for row in source_rows)),
            "exact_duplicates_removed": exact_dropped,
            "exact_conflict_groups_quarantined": len(exact_conflicts),
            "near_conflict_groups_quarantined": len(near_conflicts),
            "family_representatives_before_overlap": len(representatives),
            "exact_overlaps_removed": exact_overlaps_removed,
            "near_overlaps_removed": near_overlaps_removed,
            "final_rows": len(candidates),
            "final_labels": dict(Counter(str(row["label"]) for row in candidates)),
        },
        "near_template_stats": near_stats,
        "artifact": {"path": str(artifact), "sha256": sha256(artifact)},
    }
    (output / "manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=Path("data/raw/chichewa_sms_fraud.xlsx"))
    parser.add_argument("--data", type=Path, default=Path("data/processed"))
    parser.add_argument("--output", type=Path, default=Path("data/external/chichewa"))
    args = parser.parse_args()
    if not args.source.is_file():
        raise FileNotFoundError(
            f"missing Chichewa source; run scripts/fetch_datasets.py: {args.source}"
        )
    print(json.dumps(build(args.source, args.data, args.output), indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
